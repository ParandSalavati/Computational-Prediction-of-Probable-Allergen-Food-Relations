import pandas as pd
import networkx as nx
from networkx.algorithms import bipartite
from openpyxl import Workbook

# Load data from an Excel file
df = pd.read_excel('Top10Allergens.xlsx')

# Create a bipartite graph from the DataFrame
B = nx.Graph()
B.add_nodes_from(df['Common name'].unique(), bipartite=0) # Add one set of nodes
B.add_nodes_from(df['Keyword'].unique(), bipartite=1) # Add the other set of nodes
B.add_edges_from([(row['Common name'], row['Keyword']) for idx, row in df.iterrows()]) # Add edges

# Get the two sets of nodes
top_nodes = {n for n, d in B.nodes(data=True) if d['bipartite']==0}
bottom_nodes = set(B) - top_nodes

# Compute the bipartite projection of the graph
G_top = nx.bipartite.projected_graph(B, top_nodes)
G_bottom = nx.bipartite.projected_graph(B, bottom_nodes)

# Calculate Common Neighbors (CN)
cn_top = [(e[0], e[1], len(list(nx.common_neighbors(G_top, e[0], e[1])))) for e in nx.non_edges(G_top)]
cn_bottom = [(e[0], e[1], len(list(nx.common_neighbors(G_bottom, e[0], e[1])))) for e in nx.non_edges(G_bottom)]

# Calculate Jaccard Coefficient (JC)
jc_top = list(nx.jaccard_coefficient(G_top))
jc_bottom = list(nx.jaccard_coefficient(G_bottom))

# Calculate Preferential AttachmentPA)
pa_top = list(nx.preferential_attachment(G_top))
pa_bottom = list(nx.preferential_attachment(G_bottom))

# Calculate Adamic-Adar (AA)
aa_top = list(nx.adamic_adar_index(G_top))
aa_bottom = list(nx.adamic_adar_index(G_bottom))

# an workbookworkbook = Workbook()
workbook = Workbook()

# Create separate sheets for each chart
cn_top_sheet = workbook.create_sheet(title='CN (Top)')
cn_bottom_sheet = workbook.create_sheet(title="CN (Bottom)")
jc_top_sheet = workbook.create_sheet(title="JC (Top)")
jc_bottom_sheet = workbook.create_sheet(title="JC (Bottom)")
pa_top_sheet = workbook.create_sheet(title="PA (Top)")
pa_bottom_sheet = workbook.create_sheet(title="PA (Bottom)")
aa_top_sheet = workbook.create_sheet(title="AA (Top)")
aa_bottom_sheet = workbook.create_sheet(title="AA (Bottom)")

# Write the results to the corresponding sheets
cn_top_sheet.append(["Node 1", "Node 2", "Common Neighbors"])
cn_top_sheet.append(["---", "---", "---"])
for result in cn_top:
    cn_top_sheet.append(result)

cn_bottom_sheet.append(["Node 1", "Node 2", "Common Neighbors"])
cn_bottom_sheet.append(["---", "---", "---"])
for result in cn_bottom:
    cn_bottom_sheet.append(result)

jc_top_sheet.append(["Node 1", "Node 2", "Jaccard Coefficient"])
jc_top_sheet.append(["---", "---", "---"])
for result in jc_top:
    jc_top_sheet.append(result)

jc_bottom_sheet.append(["Node 1", "Node 2", "Jaccard Coefficient"])
jc_bottom_sheet.append(["---", "---", "---"])
for result in jc_bottom:
    jc_bottom_sheet.append(result)

pa_top_sheet.append(["Node 1", "Node 2", "Preferential Attachment"])
pa_top_sheet.append(["---", "---", "---"])
for result in pa_top:
    pa_top_sheet.append(result)

pa_bottom_sheet.append(["Node 1", "Node 2", "Preferential Attachment"])
pa_bottom_sheet.append(["---", "---", "---"])
for result in pa_bottom:
    pa_bottom_sheet.append(result)

aa_top_sheet.append(["Node 1", "Node 2", "Adamic-Adar"])
aa_top_sheet.append(["---", "---", "---"])
for result in aa_top:
    aa_top_sheet.append(result)

aa_bottom_sheet.append(["Node 1", "Node 2", "Adamic-Adar"])
aa_bottom_sheet.append(["---", "---", "---"])
for result in aa_bottom:
    aa_bottom_sheet.append(result)

# Save the Excel file
workbook.save("Results.xlsx")
